import PySimpleGUI as sg
import openpyxl as kk

sg.theme('GrayGrayGray')

def copy(r,c,b):
        ws.cell(r+2,b).value=ws.cell(r,c).value
        ws.cell(r+2,b)._style=ws.cell(r,c)._style

        ws.cell(r+3,b).value=ws.cell(r+1,c).value
        ws.cell(r+3,b)._style=ws.cell(r+1,c)._style

        ws.cell(r,c).value=""
        ws.cell(r,c)._style=None

        ws.cell(r+1,c).value=""
        ws.cell(r+1,c)._style=None
        
layout = [
    [sg.Text('文件位置'),sg.In(key = '-IN-'),sg.FileBrowse(button_text = "选择文件",target = '-IN-')],
    [sg.Checkbox('多行',key='1')],
    [sg.B('确认'),sg.B('取消')],
    [sg.B('关于作者')]
]

window = sg.Window('成绩条分割 v1.0', layout)

while True:
    event, values = window.read()
    
    if event == sg.WINDOW_CLOSED or event == '取消':

        break
    
    if event == '关于作者':
        sg.popup("Github:https://github.com/shixiangxi")
        
    if event =='确认' and values['1']==False:

        in_path=str(values['-IN-'])
        wb=kk.load_workbook(in_path)
        ws=wb.active

        for r in range(ws.max_row,2,-1):
            ws.insert_rows(r,2)
            for c in range(1,ws.max_column):
                ws.cell(r+1,c).value=ws.cell(1,c).value  
        
        out_path=sg.popup_get_file(
            '保存至……',
            save_as=True,
            default_extension='xlsx',
            file_types=(('表格文件','.xlsx'))
        )

        wb.save(filename=out_path)
        sg.popup("完成……正在退出")
        break

    if event =='确认' and values['1']==True:

        in_path=str(values['-IN-'])
        wb=kk.load_workbook(in_path)
        ws=wb.active
        a=ws.max_row

        for r in range(ws.max_row,2,-1):
            ws.insert_rows(r,4)
            for c in range(1,ws.max_column):
                ws.cell(r+3,c).value=ws.cell(1,c).value
        
        column=int(sg.popup_get_text('从第几列开分'))
        

        for r in range(1,ws.max_row+1,5):
            b=1
            for c in range(column,ws.max_column+1):
                copy(r,c,b)
                b=b+1
        
        out_path=sg.popup_get_file(
            '保存至……',
            save_as=True,
            default_extension='xlsx',
            file_types=(('表格文件','.xlsx'))
        )

        wb.save(filename=out_path)
        sg.popup("完成……正在退出")
        break

window.close()